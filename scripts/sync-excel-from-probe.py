# -*- coding: utf-8 -*-
"""Sync ES接口清单.xlsx SAP section + MCP statuses from Probe_Latest.json."""
import json
import openpyxl
from pathlib import Path

ROOT = Path(r"E:\00 - 中数通ES环境\ES 接口")
PROBE = ROOT / "Probe_Latest.json"
XLSX = ROOT / "ES接口清单.xlsx"

MCP_OK = {
    "get_sales_order_status": ("OK(200)", "SAP_COM_0109 SalesOrder V4"),
    "get_product": ("OK(200)", "API_PRODUCT_SRV / Product V4"),
    "get_business_partner": ("OK(200)", "API_BUSINESS_PARTNER Customer/Supplier"),
    "get_purchase_order": ("OK(200)", "PO V2/V4 api_purchaseorder_2"),
    "get_material_stock": ("OK(200)", "API_MATERIAL_STOCK_SRV"),
    "get_supplier_invoice": ("OK(200)", "V2三条URL见SAP区；V4需SAP_COM_0054"),
    "get_cost_center": ("OK(200)", "api_cost_center A_CostCenter_2"),
}


def resp_format(protocol: str) -> str:
    return "{ value[] }" if "V4" in protocol else "{ d: { results } }"


def req_format(protocol: str) -> str:
    return "$top,$filter,sap-client=100,$format=json" if "V2" in protocol else "$top,$filter,sap-client=100,$format=json"


def main():
    probe = json.loads(PROBE.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["ES接口清单"]

    # preserve MCP (2-21) and config (last 2 rows) from current sheet
    mcp_rows = []
    for r in range(2, 22):
        mcp_rows.append([ws.cell(r, c).value for c in range(1, 11)])

    config_rows = []
    for r in range(ws.max_row - 1, ws.max_row + 1):
        config_rows.append([ws.cell(r, c).value for c in range(1, 11)])

    # update MCP connectivity
    for row in mcp_rows:
        name = row[1]
        if name in MCP_OK:
            status, note = MCP_OK[name]
            row[8] = status
            row[9] = note

    # clear data rows (keep header row 1)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # rewrite MCP
    for i, row in enumerate(mcp_rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(i, c).value = val

    # SAP / EPC / 主数据 from probe
    start = 22
    for i, item in enumerate(probe):
        r = start + i
        url = item.get("读取示例") or item["接口地址"]
        if url.startswith("GET "):
            url = url[4:]
        scene = item.get("通信场景", "")
        note = item.get("备注", "")
        if scene:
            note = f"[{scene}] {note}".strip()
        ws.cell(r, 1).value = item["分类"]
        ws.cell(r, 2).value = item["接口名称"]
        ws.cell(r, 3).value = item["方法"]
        ws.cell(r, 4).value = url
        ws.cell(r, 5).value = item["协议"]
        ws.cell(r, 6).value = req_format(item["协议"])
        ws.cell(r, 7).value = resp_format(item["协议"])
        ws.cell(r, 8).value = "Basic Auth"
        ws.cell(r, 9).value = item["连通性"]
        ws.cell(r, 10).value = note

    # config rows
    cfg_start = start + len(probe)
    for i, row in enumerate(config_rows):
        r = cfg_start + i
        for c, val in enumerate(row, start=1):
            ws.cell(r, c).value = val

    # update config row for user.txt format hint
    for r in range(cfg_start, cfg_start + len(config_rows)):
        if ws.cell(r, 2).value == "SAP 凭证文件":
            ws.cell(r, 6).value = "通信用户/密码（中文）或 User Name/Password"
            ws.cell(r, 9).value = "已验证 2026-06-22"
            ws.cell(r, 10).value = "示例：接口调用的通信用户：EPC_USER"

    wb.save(XLSX)
    ok = sum(1 for x in probe if str(x["连通性"]).startswith("OK"))
    print(f"Saved {XLSX}")
    print(f"MCP rows: {len(mcp_rows)} | SAP probe rows: {len(probe)} (OK {ok}) | Config: {len(config_rows)}")
    print(f"Total rows: {ws.max_row}")


if __name__ == "__main__":
    main()
